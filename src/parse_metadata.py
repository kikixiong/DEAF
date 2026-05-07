"""Step 1: Parse raw audio data into metadata CSVs.

Data layout expected:
  Task1/text_samples_EMIS.csv + Task1/audios_EMIS/  -> ESC clips
  Task2/Text.xlsx              + Task2/noisy_speech/ -> BSC clips
  Task3/SIC.xlsx               + Task3/SIC_clips/    -> SIC clips

ESC audio filename:
  {row_id}_{text_emotion}_{mention_type}_{speech_emotion}_{speaker_id}_{tts}.wav
  e.g. 01_angry_explicit_angry_0011_STYLE.wav
  For neutral text: {row_id}_neutral_{speech_emotion}_{speaker_id}_{tts}.wav
  e.g. 26_neutral_angry_0016_F5TTS.wav

BSC audio filename:
  {text_env}_{sentence_id}_{bg_env}_{snr}.wav
  e.g. DKITCHEN_E01_DLIVING_-10.wav

SIC audio filename:
  {code}__{voice_age}_{voice_gender}.wav   (double underscore separator)
  e.g. AGE_EX_EL_01__elderly_male.wav

Outputs:
  data/metadata/esc_clips.csv
  data/metadata/bsc_clips.csv
  data/metadata/sic_clips.csv
"""

import argparse
import csv
import os
from pathlib import Path


# ---------------------------------------------------------------------------
# ESC: Emotion-Semantic Conflict
# ---------------------------------------------------------------------------

EMOTIONS = ["happy", "sad", "angry", "neutral"]


def parse_esc(task1_dir: Path, out_csv: Path):
    """Parse ESC clips from Task1 directory."""
    audio_dir = task1_dir / "audios_EMIS"
    csv_path = task1_dir / "text_samples_EMIS.csv"

    if not audio_dir.exists():
        print(f"[ESC] Audio dir not found: {audio_dir} - skipping")
        return

    text_lookup = {}
    if csv_path.exists():
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, start=1):
                row_id = f"{i:02d}"
                for col in reader.fieldnames:
                    if col == "neutral":
                        text_lookup[(row_id, "neutral", "")] = row[col]
                    else:
                        parts = col.split("_", 1)
                        if len(parts) == 2:
                            text_lookup[(row_id, parts[0], parts[1])] = row[col]
        print(f"  [ESC] Loaded {len(text_lookup)} text entries from CSV")

    rows = []
    for wav in sorted(audio_dir.rglob("*.wav")):
        stem = wav.stem
        parts = stem.split("_")

        # 6 parts: row_id, text_emotion, mention_type, speech_emotion, speaker_id, tts
        # 5 parts: row_id, neutral, speech_emotion, speaker_id, tts
        if len(parts) == 6:
            row_id, text_emo, mention, speech_emo, speaker_id, tts = parts
        elif len(parts) == 5:
            row_id, text_emo, speech_emo, speaker_id, tts = parts
            mention = ""
        else:
            print(f"  [WARN] Cannot parse: {wav.name}")
            continue

        text_emo = text_emo.lower()
        speech_emo = speech_emo.lower()
        mention = mention.lower()

        is_matched = text_emo == speech_emo
        text = text_lookup.get((row_id, text_emo, mention), "")
        clip_id = f"esc_{stem}"

        rows.append({
            "clip_id": clip_id,
            "audio_path": f"audios_EMIS/{wav.name}",
            "row_id": row_id,
            "text_emotion": text_emo,
            "mention_type": mention,
            "speech_emotion": speech_emo,
            "speaker_id": speaker_id,
            "tts_model": tts,
            "is_matched": str(is_matched),
            "text": text,
        })

    fields = [
        "clip_id", "audio_path", "row_id", "text_emotion", "mention_type",
        "speech_emotion", "speaker_id", "tts_model", "is_matched", "text",
    ]
    _write_csv(out_csv, rows, fields)
    print(f"[ESC] Wrote {len(rows)} clips -> {out_csv}")
    _print_esc_summary(rows)


def _print_esc_summary(rows):
    matched = sum(1 for r in rows if r["is_matched"] == "True")
    mismatched = len(rows) - matched
    text_emos = sorted(set(r["text_emotion"] for r in rows))
    speech_emos = sorted(set(r["speech_emotion"] for r in rows))
    mentions = sorted(set(r["mention_type"] for r in rows if r["mention_type"]))
    tts_models = sorted(set(r["tts_model"] for r in rows))
    print(f"  Total: {len(rows)}  Matched: {matched}  Mismatched: {mismatched}")
    print(f"  Text emotions: {text_emos}")
    print(f"  Speech emotions: {speech_emos}")
    print(f"  Mention types: {mentions}")
    print(f"  TTS models: {tts_models}")


# ---------------------------------------------------------------------------
# BSC: Background-Sound Conflict
# ---------------------------------------------------------------------------

ENV_TO_CATEGORY = {
    "DWASHING": "domestic", "DKITCHEN": "domestic", "DLIVING": "domestic",
    "NFIELD": "nature", "NRIVER": "nature", "NPARK": "nature",
    "OOFFICE": "office", "OHALLWAY": "office", "OMEETING": "office",
    "PSTATION": "public", "PCAFETER": "public", "PRESTO": "public",
    "STRAFFIC": "street", "SPSQUARE": "street", "SCAFE": "street",
    "TMETRO": "transportation", "TBUS": "transportation", "TCAR": "transportation",
}


def parse_bsc(task2_dir: Path, out_csv: Path):
    """Parse BSC clips from Task2 directory."""
    audio_dir = task2_dir / "noisy_speech"
    xlsx_path = task2_dir / "Text.xlsx"

    if not audio_dir.exists():
        print(f"[BSC] Audio dir not found: {audio_dir} - skipping")
        return

    text_lookup = {}
    if xlsx_path.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                code = str(row[0].value).strip() if row[0].value else ""
                sentence = str(row[1].value).strip() if row[1].value else ""
                if code:
                    text_lookup[code] = sentence
            wb.close()
            print(f"  [BSC] Loaded {len(text_lookup)} sentences from xlsx")
        except ImportError:
            print("  [BSC] openpyxl not installed, no sentence text available")

    rows = []
    for wav in sorted(audio_dir.rglob("*.wav")):
        stem = wav.stem
        parts = stem.split("_")
        if len(parts) < 4:
            print(f"  [WARN] Cannot parse: {wav.name}")
            continue

        text_env = parts[0].upper()
        sentence_id = parts[1].upper()
        bg_env = parts[2].upper()
        snr = "_".join(parts[3:])  # handles negative like -10

        code = f"{text_env}_{sentence_id}"
        text = text_lookup.get(code, "")

        text_cat = ENV_TO_CATEGORY.get(text_env, "unknown")
        bg_cat = ENV_TO_CATEGORY.get(bg_env, "unknown")

        if text_env == bg_env:
            mismatch_type = "matched"
        elif text_cat == bg_cat:
            mismatch_type = "within"
        else:
            mismatch_type = "cross"

        clip_id = f"bsc_{stem}"

        rows.append({
            "clip_id": clip_id,
            "audio_path": f"noisy_speech/{wav.name}",
            "text_env": text_env,
            "sentence_id": sentence_id,
            "code": code,
            "background_env": bg_env,
            "background_category": bg_cat,
            "text_category": text_cat,
            "mismatch_type": mismatch_type,
            "snr": snr,
            "is_matched": str(mismatch_type == "matched"),
            "text": text,
        })

    fields = [
        "clip_id", "audio_path", "text_env", "sentence_id", "code",
        "background_env", "background_category", "text_category",
        "mismatch_type", "snr", "is_matched", "text",
    ]
    _write_csv(out_csv, rows, fields)
    print(f"[BSC] Wrote {len(rows)} clips -> {out_csv}")
    _print_bsc_summary(rows)


def _print_bsc_summary(rows):
    matched = sum(1 for r in rows if r["is_matched"] == "True")
    within = sum(1 for r in rows if r["mismatch_type"] == "within")
    cross = sum(1 for r in rows if r["mismatch_type"] == "cross")
    text_envs = sorted(set(r["text_env"] for r in rows))
    bg_envs = sorted(set(r["background_env"] for r in rows))
    snrs = sorted(set(r["snr"] for r in rows))
    print(f"  Total: {len(rows)}  Matched: {matched}  Within: {within}  Cross: {cross}")
    print(f"  Text envs ({len(text_envs)}): {text_envs}")
    print(f"  BG envs ({len(bg_envs)}): {bg_envs}")
    print(f"  SNRs: {snrs}")


# ---------------------------------------------------------------------------
# SIC: Speaker-Identity Conflict
# ---------------------------------------------------------------------------
# Filename: {sub_dim}_{mention}_{text_id}_{num}__{voice_age}_{voice_gender}.wav
#
# Sub-dimensions:
#   AGE: text_id = EL(elderly) / YG(young)
#   GDR: text_id = F(female) / M(male)
#   CMB: text_id = EF(elderly_female) / YF(young_female) / YM(young_male) / EM(elderly_male)
#   NEU: text_id = NA (neutral, no identity cues in text)
#
# Voice part after __: {voice_age}_{voice_gender} e.g. elderly_male

AGE_TEXT_MAP = {"EL": "elderly", "YG": "young"}
GDR_TEXT_MAP = {"F": "female", "M": "male"}
CMB_TEXT_MAP = {
    "EF": ("elderly", "female"),
    "YF": ("young", "female"),
    "YM": ("young", "male"),
    "EM": ("elderly", "male"),
}


def parse_sic(task3_dir: Path, out_csv: Path):
    """Parse SIC clips from Task3 directory."""
    audio_dir = task3_dir / "SIC_clips"
    xlsx_path = task3_dir / "SIC.xlsx"

    if not audio_dir.exists():
        print(f"[SIC] Audio dir not found: {audio_dir} - skipping")
        return

    text_lookup = {}
    if xlsx_path.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                code = str(row[0].value).strip() if row[0].value else ""
                sentence = str(row[1].value).strip() if row[1].value else ""
                if code:
                    text_lookup[code] = sentence
            wb.close()
            print(f"  [SIC] Loaded {len(text_lookup)} sentences from xlsx")
        except ImportError:
            print("  [SIC] openpyxl not installed, no sentence text available")

    rows = []
    for wav in sorted(audio_dir.rglob("*.wav")):
        stem = wav.stem

        if "__" not in stem:
            print(f"  [WARN] No __ separator in: {wav.name}")
            continue

        text_part, voice_part = stem.split("__", 1)
        voice_parts = voice_part.split("_")
        if len(voice_parts) < 2:
            print(f"  [WARN] Cannot parse voice part: {wav.name}")
            continue

        voice_age = voice_parts[0].lower()
        voice_gender = voice_parts[1].lower()

        tp = text_part.split("_")
        sub_dim = tp[0].upper() if len(tp) >= 1 else ""
        mention = tp[1].upper() if len(tp) >= 2 else ""
        text_id = tp[2].upper() if len(tp) >= 3 else ""
        num = tp[3] if len(tp) >= 4 else ""

        semantic_age = ""
        semantic_gender = ""

        if sub_dim == "AGE":
            semantic_age = AGE_TEXT_MAP.get(text_id, "")
        elif sub_dim == "GDR":
            semantic_gender = GDR_TEXT_MAP.get(text_id, "")
        elif sub_dim == "CMB":
            sa, sg = CMB_TEXT_MAP.get(text_id, ("", ""))
            semantic_age = sa
            semantic_gender = sg
        elif sub_dim == "NEU":
            pass

        if sub_dim == "GDR":
            is_matched = semantic_gender == voice_gender
        elif sub_dim == "AGE":
            is_matched = semantic_age == voice_age
        elif sub_dim == "CMB":
            is_matched = (semantic_age == voice_age) and (semantic_gender == voice_gender)
        elif sub_dim == "NEU":
            is_matched = True
        else:
            is_matched = False

        text = text_lookup.get(text_part, "")
        clip_id = f"sic_{stem}"

        rows.append({
            "clip_id": clip_id,
            "audio_path": f"SIC_clips/{wav.name}",
            "code": text_part,
            "sub_dimension": sub_dim,
            "mention_type": mention,
            "text_id": text_id,
            "voice_age": voice_age,
            "voice_gender": voice_gender,
            "semantic_age": semantic_age,
            "semantic_gender": semantic_gender,
            "is_matched": str(is_matched),
            "text": text,
        })

    fields = [
        "clip_id", "audio_path", "code", "sub_dimension", "mention_type",
        "text_id", "voice_age", "voice_gender", "semantic_age",
        "semantic_gender", "is_matched", "text",
    ]
    _write_csv(out_csv, rows, fields)
    print(f"[SIC] Wrote {len(rows)} clips -> {out_csv}")
    _print_sic_summary(rows)


def _print_sic_summary(rows):
    matched = sum(1 for r in rows if r["is_matched"] == "True")
    mismatched = len(rows) - matched
    dims = sorted(set(r["sub_dimension"] for r in rows))
    mentions = sorted(set(r["mention_type"] for r in rows))
    v_genders = sorted(set(r["voice_gender"] for r in rows))
    v_ages = sorted(set(r["voice_age"] for r in rows))
    print(f"  Total: {len(rows)}  Matched: {matched}  Mismatched: {mismatched}")
    print(f"  Sub-dimensions: {dims}")
    print(f"  Mention types: {mentions}")
    print(f"  Voice genders: {v_genders}")
    print(f"  Voice ages: {v_ages}")

    from collections import Counter
    dim_counts = Counter(r["sub_dimension"] for r in rows)
    for d, cnt in sorted(dim_counts.items()):
        m = sum(1 for r in rows if r["sub_dimension"] == d and r["is_matched"] == "True")
        print(f"    {d}: {cnt} clips ({m} matched, {cnt-m} mismatched)")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Parse raw audio data -> metadata CSVs")
    parser.add_argument("--data-dir", type=str,
                        default=os.environ.get("DEAF_DATA_DIR", "data/raw"),
                        help="Root data directory containing Task1/Task2/Task3")
    parser.add_argument("--output-dir", type=str, default="data/metadata",
                        help="Output directory for metadata CSVs")
    parser.add_argument("--tasks", type=str, default="1,2,3",
                        help="Comma-separated task numbers to process")
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    tasks = [t.strip() for t in args.tasks.split(",")]

    if "1" in tasks:
        task1_dir = data_dir / "Task1"
        if task1_dir.exists():
            parse_esc(task1_dir, out_dir / "esc_clips.csv")
        else:
            print(f"[ESC] Task1 dir not found: {task1_dir}")

    if "2" in tasks:
        task2_dir = data_dir / "Task2"
        if task2_dir.exists():
            parse_bsc(task2_dir, out_dir / "bsc_clips.csv")
        else:
            print(f"[BSC] Task2 dir not found: {task2_dir}")

    if "3" in tasks:
        task3_dir = data_dir / "Task3"
        if task3_dir.exists():
            parse_sic(task3_dir, out_dir / "sic_clips.csv")
        else:
            print(f"[SIC] Task3 dir not found: {task3_dir}")

    print("\nDone.")


if __name__ == "__main__":
    main()
